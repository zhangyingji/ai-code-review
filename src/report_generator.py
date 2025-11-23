"""
报告生成模块
支持多种格式的评审报告生成
"""
import os
import json
from typing import Dict, List
from datetime import datetime
from jinja2 import Template
from src.simple_html_template import SIMPLE_HTML_TEMPLATE
import logging

try:
    from openpyxl import Workbook as OpenpyxlWorkbook
    from openpyxl.styles import Font as OpenpyxlFont, PatternFill as OpenpyxlPatternFill
    from openpyxl.styles import Alignment as OpenpyxlAlignment, Border as OpenpyxlBorder
    from openpyxl.styles import Side as OpenpyxlSide
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    OpenpyxlWorkbook = None  # type: ignore
    OpenpyxlFont = None  # type: ignore
    OpenpyxlPatternFill = None  # type: ignore
    OpenpyxlAlignment = None  # type: ignore
    OpenpyxlBorder = None  # type: ignore
    OpenpyxlSide = None  # type: ignore

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ReportGenerator:
    """评审报告生成器"""
    
    # 严重程度排序映射
    SEVERITY_ORDER = {
        'critical': 0,
        'major': 1,
        'minor': 2,
        'suggestion': 3
    }
    
    def _sort_issues_by_severity(self, issues: List[Dict]) -> List[Dict]:
        """
        按严重程度排序问题（从严重到建议）
        
        Args:
            issues: 问题列表
            
        Returns:
            排序后的问题列表
        """
        return sorted(issues, key=lambda x: self.SEVERITY_ORDER.get(x.get('severity', ''), 999))
    
    def __init__(self, output_dir: str = "./reports"):
        """
        初始化报告生成器
        
        Args:
            output_dir: 报告输出目录
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_report(self, review_data: Dict, format: str = "html", 
                       group_by_author: bool = True) -> str:
        """
        生成评审报告
        
        Args:
            review_data: 评审数据
            format: 报告格式 (html, markdown, json, excel)
            group_by_author: 是否按作者分组
            
        Returns:
            报告文件路径
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        source_branch = review_data['metadata']['source_branch'].replace('/', '_')
        
        if format == "html":
            filename = f"review_{source_branch}_{timestamp}.html"
            content = self._generate_html_report(review_data, group_by_author)
        elif format == "markdown":
            filename = f"review_{source_branch}_{timestamp}.md"
            content = self._generate_markdown_report(review_data, group_by_author)
        elif format == "json":
            filename = f"review_{source_branch}_{timestamp}.json"
            content = json.dumps(review_data, indent=2, ensure_ascii=False)
        elif format == "excel":
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl 库未安装，请运行: pip install openpyxl")
            filename = f"review_{source_branch}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            self._generate_excel_report(review_data, filepath, group_by_author)
            logger.info(f"报告已生成: {filepath}")
            return filepath
        else:
            raise ValueError(f"不支持的格式: {format}")
        
        filepath = os.path.join(self.output_dir, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
        
        logger.info(f"报告已生成: {filepath}")
        return filepath
    
    def _generate_html_report(self, review_data: Dict, group_by_author: bool) -> str:
        """生成HTML格式报告 - 使用简化模板"""
        # 对所有问题进行排序（输出前）
        if group_by_author and review_data.get('author_stats'):
            for author in review_data['author_stats']:
                if author.get('issues'):
                    author['issues'] = self._sort_issues_by_severity(author['issues'])
        
        for file_review in review_data.get('file_reviews', []):
            if file_review.get('issues'):
                file_review['issues'] = self._sort_issues_by_severity(file_review['issues'])
        
        # 使用简化模板
        template = Template(SIMPLE_HTML_TEMPLATE)
        return template.render(
            review_data=review_data,
            severity_labels=SEVERITY_LABELS
        )
    
    def _generate_markdown_report(self, review_data: Dict, group_by_author: bool) -> str:
        """生成Markdown格式报告"""
        lines = []
        metadata = review_data['metadata']
        stats = review_data['statistics']
        
        # 标题和基本信息
        lines.append(f"# 代码评审报告")
        lines.append(f"\n## 基本信息\n")
        lines.append(f"- **源分支**: {metadata['source_branch']}")
        lines.append(f"- **目标分支**: {metadata['target_branch']}")
        lines.append(f"- **评审时间**: {metadata['review_time']}")
        lines.append(f"- **评审耗时**: {metadata['duration_seconds']:.2f} 秒")
        lines.append(f"- **提交数量**: {metadata['total_commits']}")
        lines.append(f"- **文件变更**: {metadata['total_files_changed']}")
        lines.append(f"- **评审文件**: {metadata['total_files_reviewed']}")
        
        # 统计信息
        lines.append(f"\n## 问题统计\n")
        lines.append(f"- **总问题数**: {stats['total_issues']}")
        lines.append(f"- **严重问题**: {stats['by_severity']['critical']}")
        lines.append(f"- **主要问题**: {stats['by_severity']['major']}")
        lines.append(f"- **次要问题**: {stats['by_severity']['minor']}")
        lines.append(f"- **建议**: {stats['by_severity']['suggestion']}")
        lines.append(f"- **代码增加**: +{stats['total_additions']} 行")
        lines.append(f"- **代码删除**: -{stats['total_deletions']} 行")
        
        # 按作者分组
        if group_by_author and review_data.get('author_stats'):
            lines.append(f"\n## 按提交人统计\n")
            for author in review_data['author_stats']:
                lines.append(f"\n### {author['name']} ({author['email']})\n")
                lines.append(f"- **提交数**: {author['commit_count']}")
                lines.append(f"- **修改文件**: {author['file_count']}")
                lines.append(f"- **问题数**: {author['issue_count']}")
                
                severity = author['issue_by_severity']
                lines.append(f"  - 严重: {severity['critical']}")
                lines.append(f"  - 主要: {severity['major']}")
                lines.append(f"  - 次要: {severity['minor']}")
                lines.append(f"  - 建议: {severity['suggestion']}")
                
                # ... existing code ...
                
                # 列出该作者相关的问题 - 优先显示严重问题
                if author['issues']:
                    # 按严重程度排序所有问题
                    sorted_issues = self._sort_issues_by_severity(author['issues'])
                    # 分离严重问题和其他问题
                    critical_issues = [i for i in sorted_issues if i['severity'] == 'critical']
                    other_issues = [i for i in sorted_issues if i['severity'] != 'critical']
                    
                    # 首先显示所有严重问题
                    if critical_issues:
                        lines.append(f"\n**🔴 严重问题** (共 {len(critical_issues)} 个):")
                        for issue in critical_issues:
                            severity_label = SEVERITY_LABELS.get(issue['severity'], issue['severity'])
                            line_info = f" (第 {issue.get('line', 'N/A')} 行)" if issue.get('line') else ""
                            lines.append(f"- [{severity_label}] {issue['description']}{line_info}")
                    
                    # 然后显示其他问题（最多10个）
                    if other_issues:
                        display_count = min(10, len(other_issues))
                        lines.append(f"\n**其他问题** (显示 {display_count} 个，共 {len(other_issues)} 个):")
                        for issue in other_issues[:10]:
                            severity_label = SEVERITY_LABELS.get(issue['severity'], issue['severity'])
                            line_info = f" (第 {issue.get('line', 'N/A')} 行)" if issue.get('line') else ""
                            lines.append(f"- [{severity_label}] {issue['description']}{line_info}")
        
        # 详细文件评审结果
        lines.append(f"\n## 文件评审详情\n")
        for file_review in review_data['file_reviews']:
            lines.append(f"\n### {file_review['file_path']}\n")
            lines.append(f"- **变更**: +{file_review['additions']} -{file_review['deletions']}")
            
            if file_review['new_file']:
                lines.append(f"- **状态**: 新文件")
            if file_review['renamed_file']:
                lines.append(f"- **状态**: 重命名")
            
            lines.append(f"\n**评审总结**: {file_review.get('summary', '无')}\n")
            
            if file_review.get('issues'):
                lines.append(f"\n**发现的问题**:\n")
                # 按严重程度排序问题
                sorted_issues = self._sort_issues_by_severity(file_review['issues'])
                for i, issue in enumerate(sorted_issues, 1):
                    severity_label = SEVERITY_LABELS.get(issue['severity'], issue['severity'])
                    lines.append(f"{i}. [{severity_label}] **{issue['category']}**")
                    lines.append(f"   - 位置: {issue.get('line', 'N/A')}")
                    lines.append(f"   - 描述: {issue['description']}")
                    if issue.get('suggestion'):
                        lines.append(f"   - 建议: {issue['suggestion']}")
                    lines.append("")
        
        return "\n".join(lines)
    
    def _generate_excel_report(self, review_data: Dict, filepath: str, group_by_author: bool) -> None:
        """生成Excel格式报告"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 库未安装")
        
        wb = OpenpyxlWorkbook()  # type: ignore
        
        # 削除openpyxl默认创建的空白 Sheet
        if wb.sheetnames and wb.sheetnames[0] == 'Sheet':
            wb.remove(wb[wb.sheetnames[0]])  # type: ignore
        header_fill = OpenpyxlPatternFill(start_color="0366D6", end_color="0366D6", fill_type="solid")  # type: ignore
        header_font = OpenpyxlFont(bold=True, color="FFFFFF", size=11)  # type: ignore
        critical_fill = OpenpyxlPatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")  # type: ignore
        major_fill = OpenpyxlPatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")  # type: ignore
        minor_fill = OpenpyxlPatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # type: ignore
        center_align = OpenpyxlAlignment(horizontal="center", vertical="center", wrap_text=True)  # type: ignore
        left_align = OpenpyxlAlignment(horizontal="left", vertical="top", wrap_text=True)  # type: ignore
        border = OpenpyxlBorder(  # type: ignore
            left=OpenpyxlSide(style='thin'),  # type: ignore
            right=OpenpyxlSide(style='thin'),  # type: ignore
            top=OpenpyxlSide(style='thin'),  # type: ignore
            bottom=OpenpyxlSide(style='thin')  # type: ignore
        )
        
        # 1. 概览页
        ws = wb.create_sheet("概览")
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
        row = 1
        ws[f'A{row}'] = "代码评审报告"
        ws[f'A{row}'].font = OpenpyxlFont(size=14, bold=True)  # type: ignore
        ws.merge_cells(f'A{row}:B{row}')
        row += 2
        
        # 基本信息
        metadata = review_data['metadata']
        info_items = [
            ("源分支", metadata['source_branch']),
            ("目标分支", metadata['target_branch']),
            ("评审时间", metadata['review_time']),
            ("评审耗时", f"{metadata['duration_seconds']:.2f} 秒"),
            ("提交数量", str(metadata['total_commits'])),
            ("文件变更", str(metadata['total_files_changed'])),
        ]
        
        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = OpenpyxlFont(bold=True)  # type: ignore
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        # 统计信息
        stats = review_data['statistics']
        ws[f'A{row}'] = "问题统计"
        ws[f'A{row}'].font = OpenpyxlFont(size=12, bold=True)  # type: ignore
        row += 1
        
        stat_items = [
            ("总问题数", str(stats['total_issues'])),
            ("严重问题", str(stats['by_severity']['critical'])),
            ("主要问题", str(stats['by_severity']['major'])),
            ("次要问题", str(stats['by_severity']['minor'])),
            ("建议", str(stats['by_severity']['suggestion'])),
            ("代码增加", f"+{stats['total_additions']}"),
            ("代码删除", f"-{stats['total_deletions']}"),
        ]
        
        for label, value in stat_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = OpenpyxlFont(bold=True)  # type: ignore
            ws[f'B{row}'] = value
            row += 1
        
        # 2. 问题详情页
        ws_issues = wb.create_sheet("问题详情")
        ws_issues.column_dimensions['A'].width = 15
        ws_issues.column_dimensions['B'].width = 30
        ws_issues.column_dimensions['C'].width = 15
        ws_issues.column_dimensions['D'].width = 15
        ws_issues.column_dimensions['E'].width = 50
        ws_issues.column_dimensions['F'].width = 50
        
        # 表头
        headers = ["严重程度", "文件", "行号", "方法", "问题描述", "改进建议"]
        for col, header in enumerate(headers, 1):
            cell = ws_issues.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row = 2
        
        # 收集所有问题
        all_issues = []
        for file_review in review_data.get('file_reviews', []):
            for issue in file_review.get('issues', []):
                issue_copy = issue.copy()
                issue_copy['file_path'] = file_review['file_path']
                all_issues.append(issue_copy)
        
        # 按严重程度排序
        all_issues = self._sort_issues_by_severity(all_issues)
        
        # 填充数据
        for issue in all_issues:
            severity = issue['severity']
            
            ws_issues.cell(row=row, column=1).value = SEVERITY_LABELS.get(severity, severity)
            ws_issues.cell(row=row, column=2).value = issue.get('file_path', 'N/A')
            ws_issues.cell(row=row, column=3).value = issue.get('line', 'N/A')
            ws_issues.cell(row=row, column=4).value = issue.get('method', 'N/A')
            ws_issues.cell(row=row, column=5).value = issue.get('description', '')
            ws_issues.cell(row=row, column=6).value = issue.get('suggestion', '')
            
            # 应用样式和边框
            for col in range(1, 7):
                cell = ws_issues.cell(row=row, column=col)
                cell.border = border
                cell.alignment = left_align
                
                # 根据严重程度填充背景色
                if severity == 'critical':
                    cell.fill = critical_fill
                elif severity == 'major':
                    cell.fill = major_fill
                elif severity == 'minor':
                    cell.fill = minor_fill
            
            row += 1
            
            # 展示代码片段
            if issue.get('code_snippet'):
                # 添加空行
                row += 1
                
                # 代码片段标题
                code_title_row = row
                ws_issues.cell(row=code_title_row, column=1).value = "代码片段:"
                ws_issues.cell(row=code_title_row, column=1).font = OpenpyxlFont(bold=True, italic=True)  # type: ignore
                ws_issues.merge_cells(f'A{code_title_row}:F{code_title_row}')
                row += 1
                
                # 展示每一行代码
                code_snippet = issue['code_snippet']
                for code_line in code_snippet.get('lines', []):
                    line_num = code_line.get('line_num', '')
                    line_type = code_line.get('type', 'context')
                    line_content = code_line.get('content', '')
                    in_range = code_line.get('in_range', False)
                    
                    # 第一列：行号
                    cell = ws_issues.cell(row=row, column=1)
                    cell.value = str(line_num)
                    cell.font = OpenpyxlFont(size=9, color="666666")  # type: ignore
                    cell.border = border
                    
                    # 第二列：代码类型标记
                    type_map = {'added': '+', 'deleted': '-', 'context': ' '}
                    cell = ws_issues.cell(row=row, column=2)
                    cell.value = type_map.get(line_type, ' ')
                    cell.border = border
                    
                    # 第三列起：代码内容
                    cell = ws_issues.cell(row=row, column=3)
                    cell.value = line_content
                    cell.border = border
                    cell.alignment = left_align
                    ws_issues.merge_cells(f'C{row}:F{row}')
                    
                    # 根据类型填充背景色
                    if in_range:
                        for col in range(1, 7):
                            cell = ws_issues.cell(row=row, column=col)
                            cell.fill = minor_fill  # 黄色高亮
                    elif line_type == 'added':
                        for col in range(1, 7):
                            ws_issues.cell(row=row, column=col).fill = OpenpyxlPatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # type: ignore
                    elif line_type == 'deleted':
                        for col in range(1, 7):
                            ws_issues.cell(row=row, column=col).fill = OpenpyxlPatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # type: ignore
                    
                    row += 1
                
                row += 1  # 添加闲置下一个问题前的空行
        
        # 3. 文件评审页
        ws_files = wb.create_sheet("文件评审")
        ws_files.column_dimensions['A'].width = 30
        ws_files.column_dimensions['B'].width = 12
        ws_files.column_dimensions['C'].width = 12
        ws_files.column_dimensions['D'].width = 50
        
        headers = ["文件路径", "增加", "删除", "评审总结"]
        for col, header in enumerate(headers, 1):
            cell = ws_files.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row = 2
        for file_review in review_data.get('file_reviews', []):
            ws_files.cell(row=row, column=1).value = file_review['file_path']
            ws_files.cell(row=row, column=2).value = file_review['additions']
            ws_files.cell(row=row, column=3).value = file_review['deletions']
            ws_files.cell(row=row, column=4).value = file_review.get('summary', '')
            
            for col in range(1, 5):
                cell = ws_files.cell(row=row, column=col)
                cell.border = border
                cell.alignment = left_align
            
            row += 1
        
        # 4. 按作者统计页（如果有）
        if group_by_author and review_data.get('author_stats'):
            ws_authors = wb.create_sheet("作者统计")
            ws_authors.column_dimensions['A'].width = 15
            ws_authors.column_dimensions['B'].width = 25
            ws_authors.column_dimensions['C'].width = 10
            ws_authors.column_dimensions['D'].width = 10
            ws_authors.column_dimensions['E'].width = 10
            ws_authors.column_dimensions['F'].width = 10
            ws_authors.column_dimensions['G'].width = 10
            ws_authors.column_dimensions['H'].width = 10
            
            headers = ["作者", "邮箱", "提交数", "文件数", "问题数", "严重", "主要", "次要"]
            for col, header in enumerate(headers, 1):
                cell = ws_authors.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            
            row = 2
            for author in review_data['author_stats']:
                ws_authors.cell(row=row, column=1).value = author['name']
                ws_authors.cell(row=row, column=2).value = author['email']
                ws_authors.cell(row=row, column=3).value = author['commit_count']
                ws_authors.cell(row=row, column=4).value = author['file_count']
                ws_authors.cell(row=row, column=5).value = author['issue_count']
                ws_authors.cell(row=row, column=6).value = author['issue_by_severity']['critical']
                ws_authors.cell(row=row, column=7).value = author['issue_by_severity']['major']
                ws_authors.cell(row=row, column=8).value = author['issue_by_severity']['minor']
                
                for col in range(1, 9):
                    cell = ws_authors.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = center_align
                
                row += 1
        
        wb.save(filepath)


# 严重程度颜色映射
SEVERITY_COLORS = {
    'critical': '#d73a4a',
    'major': '#e36209',
    'minor': '#fbca04',
    'suggestion': '#0366d6'
}

# 严重程度标签
SEVERITY_LABELS = {
    'critical': '🔴 严重',
    'major': '🟠 主要',
    'minor': '🟡 次要',
    'suggestion': '🔵 建议'
}

# HTML模板
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>代码评审报告 - {{ review_data.metadata.source_branch }}</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Helvetica, Arial, sans-serif;
            line-height: 1.6;
            color: #24292e;
            background-color: #f6f8fa;
            padding: 20px;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }
        h1 {
            color: #0366d6;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 3px solid #0366d6;
        }
        h2 {
            color: #24292e;
            margin-top: 30px;
            margin-bottom: 15px;
            padding-bottom: 8px;
            border-bottom: 1px solid #e1e4e8;
        }
        h3 {
            color: #586069;
            margin-top: 20px;
            margin-bottom: 10px;
        }
        .metadata {
            background: #f6f8fa;
            padding: 15px;
            border-radius: 6px;
            margin-bottom: 20px;
        }
        .metadata-item {
            display: inline-block;
            margin-right: 30px;
            margin-bottom: 10px;
        }
        .metadata-label {
            font-weight: 600;
            color: #586069;
        }
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }
        .stat-card {
            background: #f6f8fa;
            padding: 15px;
            border-radius: 6px;
            border-left: 4px solid #0366d6;
        }
        .stat-label {
            font-size: 0.9em;
            color: #586069;
            margin-bottom: 5px;
        }
        .stat-value {
            font-size: 1.8em;
            font-weight: 700;
            color: #24292e;
        }
        .author-card {
            background: #fff;
            border: 1px solid #e1e4e8;
            border-radius: 6px;
            padding: 20px;
            margin-bottom: 20px;
        }
        .author-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 15px;
        }
        .author-name {
            font-size: 1.2em;
            font-weight: 600;
        }
        .author-stats {
            display: flex;
            gap: 20px;
            font-size: 0.9em;
        }
        .severity-badge {
            display: inline-block;
            padding: 3px 8px;
            border-radius: 12px;
            font-size: 0.85em;
            font-weight: 600;
            color: white;
            margin-right: 8px;
        }
        .issue-item {
            background: #f6f8fa;
            border-left: 4px solid #e1e4e8;
            padding: 12px;
            margin-bottom: 10px;
            border-radius: 4px;
        }
        .issue-item.critical {
            background: #fde8eb;
            border-left-color: {{ severity_colors.critical }};
            font-weight: 500;
        }
        .issue-description {
            margin: 8px 0;
        }
        .issue-suggestion {
            margin-top: 8px;
            padding: 8px;
            background: #fff;
            border-radius: 4px;
            font-size: 0.9em;
        }
        .file-card {
            background: #fff;
            border: 1px solid #e1e4e8;
            border-radius: 6px;
            padding: 15px;
            margin-bottom: 15px;
        }
        .file-header {
            font-weight: 600;
            margin-bottom: 10px;
            color: #0366d6;
        }
        .file-stats {
            font-size: 0.9em;
            color: #586069;
            margin-bottom: 10px;
        }
        .summary {
            background: #f6f8fa;
            padding: 10px;
            border-radius: 4px;
            margin: 10px 0;
            font-style: italic;
        }
        .commit-list {
            list-style: none;
            margin: 10px 0;
        }
        .commit-item {
            padding: 5px 0;
            font-size: 0.9em;
        }
        .badge-critical { background-color: {{ severity_colors.critical }}; }
        .badge-major { background-color: {{ severity_colors.major }}; }
        .badge-minor { background-color: {{ severity_colors.minor }}; }
        .badge-suggestion { background-color: {{ severity_colors.suggestion }}; }
        
        /* 折叠功能样式 */
        .collapsible-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            cursor: pointer;
            user-select: none;
            padding: 10px 0;
            border-bottom: 1px solid #e1e4e8;
            margin-bottom: 15px;
        }
        .collapsible-header:hover {
            background-color: #f6f8fa;
            border-radius: 4px;
            padding: 10px;
            margin-bottom: 15px;
            margin-left: -10px;
            margin-right: -10px;
            padding-left: 10px;
        }
        .collapse-icon {
            display: inline-block;
            width: 20px;
            height: 20px;
            text-align: center;
            font-weight: bold;
            color: #586069;
            transition: transform 0.3s ease;
        }
        .collapse-icon.collapsed {
            transform: rotate(-90deg);
        }
        .collapsible-content {
            max-height: 10000px;
            overflow: visible;
            transition: max-height 0.3s ease, opacity 0.3s ease;
            opacity: 1;
        }
        .collapsible-content.collapsed {
            max-height: 0;
            overflow: hidden;
            opacity: 0;
        }
        .dashboard {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            border-radius: 8px;
            margin-bottom: 30px;
        }
        .dashboard h2 {
            color: white;
            border-bottom: 2px solid rgba(255,255,255,0.3);
        }
        .dashboard-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 20px;
            margin-top: 20px;
        }
        .dashboard-item {
            text-align: center;
            padding: 15px;
            background: rgba(255,255,255,0.1);
            border-radius: 6px;
            backdrop-filter: blur(10px);
        }
        .dashboard-item-label {
            font-size: 0.9em;
            opacity: 0.9;
            margin-bottom: 8px;
        }
        .dashboard-item-value {
            font-size: 2em;
            font-weight: bold;
        }
        .quick-nav {
            position: fixed;
            bottom: 30px;
            right: 30px;
            background: white;
            border: 2px solid #0366d6;
            padding: 15px;
            border-radius: 8px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            z-index: 1000;
            max-width: 250px;
        }
        .quick-nav-title {
            font-weight: 600;
            margin-bottom: 10px;
            color: #24292e;
            font-size: 0.9em;
        }
        .quick-nav-button {
            display: block;
            padding: 8px 12px;
            margin-bottom: 6px;
            background: #0366d6;
            color: white;
            border-radius: 4px;
            cursor: pointer;
            font-size: 0.85em;
            text-decoration: none;
            transition: background 0.2s;
            border: none;
            width: 100%;
            text-align: left;
        }
        .quick-nav-button:hover {
            background: #0256c7;
        }
        .quick-nav-button:last-child {
            margin-bottom: 0;
        }
        .critical-issues-section {
            background: #fff5f5;
            border: 2px solid #d73a4a;
            border-radius: 8px;
            padding: 20px;
            margin-bottom: 30px;
        }
        .critical-issues-section h2 {
            color: #d73a4a;
            border-bottom: 2px solid #d73a4a;
            padding-bottom: 10px;
            margin-bottom: 15px;
        }
        .critical-issues-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(320px, 1fr));
            gap: 15px;
            margin-top: 15px;
        }
        .critical-issue-card {
            background: white;
            border-left: 4px solid #d73a4a;
            padding: 15px;
            border-radius: 4px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }
        .critical-issue-author {
            font-weight: 600;
            color: #24292e;
            font-size: 0.95em;
            margin-bottom: 5px;
        }
        .critical-issue-file {
            font-size: 0.85em;
            color: #586069;
            word-break: break-all;
        }
        .critical-issue-line {
            display: inline-block;
            padding: 2px 6px;
            background: #f6f8fa;
            border-radius: 3px;
            font-size: 0.75em;
            color: #586069;
            margin-left: 5px;
        }
        
        /* 代码段落样式 */
        .code-snippet {
            background: #f6f8fa;
            border: 1px solid #d1d5da;
            border-radius: 6px;
            margin: 10px 0;
            font-family: 'Courier New', monospace;
            font-size: 0.85em;
            overflow-x: auto;
        }
        .code-snippet-header {
            background: #f3f3f3;
            padding: 8px 12px;
            border-bottom: 1px solid #d1d5da;
            font-weight: 600;
            color: #24292e;
            cursor: pointer;
            user-select: none;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .code-snippet-header:hover {
            background: #e8e8e8;
        }
        .code-snippet-toggle {
            display: inline-block;
            width: 20px;
            height: 20px;
            text-align: center;
            transition: transform 0.3s ease;
        }
        .code-snippet-toggle.collapsed {
            transform: rotate(-90deg);
        }
        .code-snippet-content {
            max-height: 400px;
            overflow-y: auto;
            transition: max-height 0.3s ease;
        }
        .code-snippet-content.collapsed {
            max-height: 0;
            overflow: hidden;
        }
        .code-line {
            display: flex;
            padding: 2px 0;
            line-height: 1.5;
        }
        .code-line-num {
            width: 50px;
            text-align: right;
            padding-right: 12px;
            color: #586069;
            background: #f6f8fa;
            user-select: none;
            border-right: 1px solid #d1d5da;
            flex-shrink: 0;
        }
        .code-line-content {
            flex: 1;
            padding: 0 12px;
            white-space: pre-wrap;
            word-wrap: break-word;
            color: #24292e;
        }
        .code-line.added {
            background: #f0f9ff;
        }
        .code-line.added .code-line-num {
            background: #cce5ff;
        }
        .code-line.added .code-line-content {
            color: #0366d6;
        }
        .code-line.deleted {
            background: #fef2f2;
        }
        .code-line.deleted .code-line-num {
            background: #ffd7d7;
        }
        .code-line.deleted .code-line-content {
            color: #cb2431;
        }
        .code-line.in-range {
            background-color: #fff3cd !important;
        }
        .code-line.in-range .code-line-num {
            background-color: #ffe5a1 !important;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>🔍 代码评审报告</h1>
        
        <!-- 基本元数据 -->
        <div class="metadata">
            <div class="metadata-item">
                <span class="metadata-label">源分支:</span> {{ review_data.metadata.source_branch }}
            </div>
            <div class="metadata-item">
                <span class="metadata-label">目标分支:</span> {{ review_data.metadata.target_branch }}
            </div>
            <div class="metadata-item">
                <span class="metadata-label">评审时间:</span> {{ review_data.metadata.review_time }}
            </div>
            <div class="metadata-item">
                <span class="metadata-label">评审耗时:</span> {{ "%.2f"|format(review_data.metadata.duration_seconds) }} 秒
            </div>
        </div>
        
        <!-- 仪表盘 -->
        <div class="dashboard">
            <h2>📈 关键指标概览</h2>
            <div class="dashboard-grid">
                <div class="dashboard-item" onclick="filterIssuesBySeverity('all')" style="cursor: pointer; border: 2px solid transparent;" onmouseover="this.style.borderColor='#0366d6';" onmouseout="this.style.borderColor='transparent';">
                    <div class="dashboard-item-label">总问题数</div>
                    <div class="dashboard-item-value">{{ review_data.statistics.total_issues }}</div>
                </div>
                <div class="dashboard-item" onclick="filterIssuesBySeverity('critical')" style="cursor: pointer; border: 2px solid transparent;" onmouseover="this.style.borderColor='#ff6b6b';" onmouseout="this.style.borderColor='transparent';">
                    <div class="dashboard-item-label">严重问题</div>
                    <div class="dashboard-item-value" style="color: #ff6b6b;">{{ review_data.statistics.by_severity.critical }}</div>
                </div>
                <div class="dashboard-item" onclick="filterIssuesBySeverity('major')" style="cursor: pointer; border: 2px solid transparent;" onmouseover="this.style.borderColor='#ffa500';" onmouseout="this.style.borderColor='transparent';">
                    <div class="dashboard-item-label">主要问题</div>
                    <div class="dashboard-item-value" style="color: #ffa500;">{{ review_data.statistics.by_severity.major }}</div>
                </div>
                <div class="dashboard-item" onclick="filterIssuesBySeverity('minor')" style="cursor: pointer; border: 2px solid transparent;" onmouseover="this.style.borderColor='#ffd700';" onmouseout="this.style.borderColor='transparent';">
                    <div class="dashboard-item-label">次要问题</div>
                    <div class="dashboard-item-value" style="color: #ffd700;">{{ review_data.statistics.by_severity.minor }}</div>
                </div>
                <div class="dashboard-item" onclick="filterIssuesBySeverity('suggestion')" style="cursor: pointer; border: 2px solid transparent;" onmouseover="this.style.borderColor='#87ceeb';" onmouseout="this.style.borderColor='transparent';">
                    <div class="dashboard-item-label">建议</div>
                    <div class="dashboard-item-value" style="color: #87ceeb;">{{ review_data.statistics.by_severity.suggestion }}</div>
                </div>
                <div class="dashboard-item">
                    <div class="dashboard-item-label">涉及提交</div>
                    <div class="dashboard-item-value">{{ review_data.metadata.total_commits }}</div>
                </div>
            </div>
        </div>
        
        <!-- 快速导航 -->
        <div class="quick-nav">
            <div class="quick-nav-title">🎯 快速导航</div>
            {% if group_by_author and review_data.author_stats %}
                <a class="quick-nav-button" onclick="expandAllAuthors()">📂 展开所有提交人</a>
                <a class="quick-nav-button" onclick="collapseAllAuthors()">📁 折叠所有提交人</a>
            {% endif %}
            <a class="quick-nav-button" onclick="expandAllFiles()">📂 展开所有文件</a>
            <a class="quick-nav-button" onclick="collapseAllFiles()">📁 折叠所有文件</a>
        </div>
        
        <!-- 严重问题单独展示 -->
        {% if review_data.statistics.by_severity.critical > 0 %}
        <div class="critical-issues-section">
            <h2>🔴 严重问题汇总 (共 {{ review_data.statistics.by_severity.critical }} 个)</h2>
            <div class="critical-issues-grid">
            {% for author in review_data.author_stats %}
                {% set author_critical = author.issues | selectattr('severity', 'equalto', 'critical') | list %}
                {% for issue in author_critical %}
                <div class="critical-issue-card">
                    <div class="critical-issue-author">👤 {{ author.name }}</div>
                    <div class="critical-issue-file">
                        📄 {{ issue.file_path if issue.file_path else '(文件信息)' }}
                        {% if issue.line %}<span class="critical-issue-line">@ {{ issue.line }}</span>{% endif %}
                    </div>
                    <div class="issue-description" style="margin-top: 8px;">{{ issue.description }}</div>
                    {% if issue.suggestion %}
                    <div class="issue-suggestion" style="margin-top: 8px;">
                        💡 <strong>建议:</strong> {{ issue.suggestion }}
                    </div>
                    {% endif %}
                </div>
                {% endfor %}
            {% endfor %}
            </div>
        </div>
        {% endif %}
        
        {% if group_by_author and review_data.author_stats %}
        <h2 class="collapsible-header" onclick="toggleAuthorSection(this)">
            <span>👥 按提交人统计</span>
            <span class="collapse-icon">▼</span>
        </h2>
        <div class="author-section collapsible-content">
        {% for author in review_data.author_stats %}
        <div class="author-card">
            <div class="collapsible-header" onclick="toggleAuthorCard(this)">
                <div>
                    <div class="author-name">{{ author.name }}
                        <span style="color: #586069; font-size: 0.8em; font-weight: normal;" class="author-stats">
                            📝 {{ author.commit_count }} | 📁 {{ author.file_count }} | ⚠️ {{ author.issue_count }} 个问题
                        </span>
                    </div>
                </div>
                <span class="collapse-icon">▼</span>
            </div>
            
            <div class="collapsible-content">
                <div style="margin-bottom: 15px;">
                    <span class="severity-badge badge-critical">严重 {{ author.issue_by_severity.critical }}</span>
                    <span class="severity-badge badge-major">主要 {{ author.issue_by_severity.major }}</span>
                    <span class="severity-badge badge-minor">次要 {{ author.issue_by_severity.minor }}</span>
                    <span class="severity-badge badge-suggestion">建议 {{ author.issue_by_severity.suggestion }}</span>
                </div>
                
                {% if author.commits %}
                <h3>最近提交</h3>
                <ul class="commit-list">
                    {% for commit in author.commits[:5] %}
                    <li class="commit-item">
                        <code>{{ commit.short_id }}</code> {{ commit.title }}
                    </li>
                    {% endfor %}
                </ul>
                {% endif %}
                
                {% if author.issues %}
                {% set critical_issues = author.issues | selectattr('severity', 'equalto', 'critical') | list %}
                {% set other_issues = author.issues | rejectattr('severity', 'equalto', 'critical') | list %}
                
                {% if critical_issues %}
                <h3>🔴 严重问题 (共 {{ critical_issues|length }} 个)</h3>
                {% for issue in critical_issues %}
                <div class="issue-item critical" style="border-left-color: {{ severity_colors[issue.severity] }}">
                    <div>
                        <span class="severity-badge badge-{{ issue.severity }}">
                            {{ severity_labels[issue.severity] }}
                        </span>
                        <strong>{{ issue.category }}</strong>
                    </div>
                    {% if issue.file_path %}
                    <div style="margin-top: 8px; padding: 8px; background: #f6f8fa; border-radius: 4px; font-size: 0.9em;">
                        <div><strong>📋 文件:</strong> {{ issue.file_path }}</div>
                        <div><strong>🔍 位置:</strong> 第 {{ issue.line }} 行{% if issue.method %} - 方法: <code>{{ issue.method }}</code>{% endif %}</div>
                    </div>
                    {% else %}
                    <div style="margin-top: 8px; padding: 8px; background: #f6f8fa; border-radius: 4px; font-size: 0.9em;">
                        <div><strong>🔍 位置:</strong> 第 {{ issue.line }} 行{% if issue.method %} - 方法: <code>{{ issue.method }}</code>{% endif %}</div>
                    </div>
                    {% endif %}
                    <div class="issue-description" style="margin-top: 8px;">{{ issue.description }}</div>
                    {% if issue.suggestion %}
                    <div class="issue-suggestion">
                        💡 <strong>建议:</strong> {{ issue.suggestion }}
                    </div>
                    {% endif %}
                    
                    {% if issue.code_snippet %}
                    <div class="code-snippet" style="margin-top: 8px;">
                        <div class="code-snippet-header" onclick="toggleCodeSnippet(this)">
                            <span>{{ issue.code_snippet.start_line }}-{{ issue.code_snippet.end_line }} 行 的代码段落</span>
                            <span class="code-snippet-toggle">\u25bc</span>
                        </div>
                        <div class="code-snippet-content">
                            {% for line in issue.code_snippet.lines %}
                            <div class="code-line {% if line.type %}{{ line.type }}{% endif %}{% if line.in_range %} in-range{% endif %}">
                                <div class="code-line-num">{{ line.line_num }}</div>
                                <div class="code-line-content">{{ line.content }}</div>
                            </div>
                            {% endfor %}
                        </div>
                    </div>
                    {% endif %}
                </div>
                {% endfor %}
                {% endif %}
                
                {% if other_issues %}
                <h3>其他问题 (显示 {{ [10, other_issues|length]|min }} 个，共 {{ other_issues|length }} 个)</h3>
                {% for issue in other_issues[:10] %}
                <div class="issue-item" style="border-left-color: {{ severity_colors[issue.severity] }}">
                    <div>
                        <span class="severity-badge badge-{{ issue.severity }}">
                            {{ severity_labels[issue.severity] }}
                        </span>
                        <strong>{{ issue.category }}</strong>
                    </div>
                    {% if issue.file_path %}
                    <div style="margin-top: 8px; padding: 8px; background: #f6f8fa; border-radius: 4px; font-size: 0.9em;">
                        <div><strong>📋 文件:</strong> {{ issue.file_path }}</div>
                        <div><strong>🔍 位置:</strong> 第 {{ issue.line }} 行{% if issue.method %} - 方法: <code>{{ issue.method }}</code>{% endif %}</div>
                    </div>
                    {% else %}
                    <div style="margin-top: 8px; padding: 8px; background: #f6f8fa; border-radius: 4px; font-size: 0.9em;">
                        <div><strong>🔍 位置:</strong> 第 {{ issue.line }} 行{% if issue.method %} - 方法: <code>{{ issue.method }}</code>{% endif %}</div>
                    </div>
                    {% endif %}
                    <div class="issue-description" style="margin-top: 8px;">{{ issue.description }}</div>
                    {% if issue.suggestion %}
                    <div class="issue-suggestion">
                        💡 <strong>建议:</strong> {{ issue.suggestion }}
                    </div>
                    {% endif %}
                    
                    {% if issue.code_snippet %}
                    <div class="code-snippet" style="margin-top: 8px;">
                        <div class="code-snippet-header" onclick="toggleCodeSnippet(this)">
                            <span>{{ issue.code_snippet.start_line }}-{{ issue.code_snippet.end_line }} 行 的代码段落</span>
                            <span class="code-snippet-toggle">\u25bc</span>
                        </div>
                        <div class="code-snippet-content">
                            {% for line in issue.code_snippet.lines %}
                            <div class="code-line {% if line.type %}{{ line.type }}{% endif %}{% if line.in_range %} in-range{% endif %}">
                                <div class="code-line-num">{{ line.line_num }}</div>
                                <div class="code-line-content">{{ line.content }}</div>
                            </div>
                            {% endfor %}
                        </div>
                    </div>
                    {% endif %}
                </div>
                {% endfor %}
                {% endif %}
                {% endif %}
            </div>
        </div>
        {% endfor %}
        </div>
        {% endif %}
        
        <h2 class="collapsible-header" onclick="toggleFileSection(this)">
            <span>📝 文件评审详情</span>
            <span class="collapse-icon">▼</span>
        </h2>
        <div class="file-section collapsible-content">
        {% for file_review in review_data.file_reviews %}
        <div class="file-card">
            <div class="collapsible-header" onclick="toggleFileCard(this)">
                <div style="flex-grow: 1;">
                    <div class="file-header">{{ file_review.file_path }}</div>
                    <div class="file-stats">
                        <span style="color: #28a745;">+{{ file_review.additions }}</span>
                        <span style="color: #d73a4a;">-{{ file_review.deletions }}</span>
                        {% if file_review.new_file %}<span> | 新文件</span>{% endif %}
                        {% if file_review.renamed_file %}<span> | 重命名</span>{% endif %}
                    </div>
                </div>
                <span class="collapse-icon">▼</span>
            </div>
            
            <div class="collapsible-content">
                <div class="summary">
                    <strong>评审总结:</strong> {{ file_review.summary }}
                </div>
                
                {% if file_review.issues %}
                <h3>发现的问题</h3>
                {% for issue in file_review.issues %}
                <div class="issue-item" style="border-left-color: {{ severity_colors[issue.severity] }}">
                    <div>
                        <span class="severity-badge badge-{{ issue.severity }}">
                            {{ severity_labels[issue.severity] }}
                        </span>
                        <strong>{{ issue.category }}</strong>
                    </div>
                    <div style="margin-top: 8px; padding: 8px; background: #f6f8fa; border-radius: 4px; font-size: 0.9em;">
                        <div><strong>📋 位置:</strong> 第 {{ issue.line }} 行{% if issue.method %} - 方法: <code>{{ issue.method }}</code>{% endif %}</div>
                    </div>
                    <div class="issue-description" style="margin-top: 8px;">{{ issue.description }}</div>
                    {% if issue.suggestion %}
                    <div class="issue-suggestion">
                        💡 <strong>改进建议:</strong> {{ issue.suggestion }}
                    </div>
                    {% endif %}
                    
                    {% if issue.code_snippet %}
                    <div class="code-snippet" style="margin-top: 8px;">
                        <div class="code-snippet-header" onclick="toggleCodeSnippet(this)">
                            <span>{{ issue.code_snippet.start_line }}-{{ issue.code_snippet.end_line }} 行 的代码段落</span>
                            <span class="code-snippet-toggle">\u25bc</span>
                        </div>
                        <div class="code-snippet-content">
                            {% for line in issue.code_snippet.lines %}
                            <div class="code-line {% if line.type %}{{ line.type }}{% endif %}{% if line.in_range %} in-range{% endif %}">
                                <div class="code-line-num">{{ line.line_num }}</div>
                                <div class="code-line-content">{{ line.content }}</div>
                            </div>
                            {% endfor %}
                        </div>
                    </div>
                    {% endif %}
                </div>
                {% endfor %}
                {% endif %}
            </div>
        </div>
        {% endfor %}
        </div>
        
        <div style="margin-top: 40px; padding-top: 20px; border-top: 1px solid #e1e4e8; text-align: center; color: #586069; font-size: 0.9em;">
            Generated by Code Review System | {{ review_data.metadata.review_time }}
        </div>
    </div>
    
    <script>
        // 折叠/展开整个作者段落
        function toggleAuthorSection(element) {
            const section = element.parentElement.querySelector('.author-section');
            const icon = element.querySelector('.collapse-icon');
            if (section.classList.contains('collapsed')) {
                section.classList.remove('collapsed');
                icon.classList.remove('collapsed');
            } else {
                section.classList.add('collapsed');
                icon.classList.add('collapsed');
            }
        }
        
        // 折叠/展开单个提交人信息板
        function toggleAuthorCard(element) {
            const content = element.nextElementSibling;
            const icon = element.querySelector('.collapse-icon');
            if (content && content.classList.contains('collapsible-content')) {
                if (content.classList.contains('collapsed')) {
                    content.classList.remove('collapsed');
                    icon.classList.remove('collapsed');
                } else {
                    content.classList.add('collapsed');
                    icon.classList.add('collapsed');
                }
            }
        }
        
        // 折叠/展开整个文件段落
        function toggleFileSection(element) {
            const section = element.parentElement.querySelector('.file-section');
            const icon = element.querySelector('.collapse-icon');
            if (section.classList.contains('collapsed')) {
                section.classList.remove('collapsed');
                icon.classList.remove('collapsed');
            } else {
                section.classList.add('collapsed');
                icon.classList.add('collapsed');
            }
        }
        
        // 折叠/展开单个文件信息板
        function toggleFileCard(element) {
            const content = element.nextElementSibling;
            const icon = element.querySelector('.collapse-icon');
            if (content && content.classList.contains('collapsible-content')) {
                if (content.classList.contains('collapsed')) {
                    content.classList.remove('collapsed');
                    icon.classList.remove('collapsed');
                } else {
                    content.classList.add('collapsed');
                    icon.classList.add('collapsed');
                }
            }
        }
        
        // 展开所有提交人
        function expandAllAuthors() {
            const section = document.querySelector('.author-section');
            if (section) section.classList.remove('collapsed');
            const headerIcon = document.querySelector('h2 .collapse-icon');
            if (headerIcon) headerIcon.classList.remove('collapsed');
            document.querySelectorAll('.author-card .collapsible-content').forEach(el => {
                el.classList.remove('collapsed');
                el.parentElement.querySelector('.collapse-icon').classList.remove('collapsed');
            });
        }
        
        // 折叠所有提交人
        function collapseAllAuthors() {
            const section = document.querySelector('.author-section');
            if (section) section.classList.add('collapsed');
            const headerIcon = document.querySelector('h2 .collapse-icon');
            if (headerIcon) headerIcon.classList.add('collapsed');
            document.querySelectorAll('.author-card .collapsible-content').forEach(el => {
                el.classList.add('collapsed');
                el.parentElement.querySelector('.collapse-icon').classList.add('collapsed');
            });
        }
        
        // 展开所有文件
        function expandAllFiles() {
            const section = document.querySelector('.file-section');
            if (section) section.classList.remove('collapsed');
            const headerIcon = document.querySelectorAll('h2 .collapse-icon')[1];
            if (headerIcon) headerIcon.classList.remove('collapsed');
            document.querySelectorAll('.file-card .collapsible-content').forEach(el => {
                el.classList.remove('collapsed');
                el.parentElement.querySelector('.collapse-icon').classList.remove('collapsed');
            });
        }
        
        // 折厠所有文件
        function collapseAllFiles() {
            const section = document.querySelector('.file-section');
            if (section) section.classList.add('collapsed');
            const headerIcon = document.querySelectorAll('h2 .collapse-icon')[1];
            if (headerIcon) headerIcon.classList.add('collapsed');
            document.querySelectorAll('.file-card .collapsible-content').forEach(el => {
                el.classList.add('collapsed');
                el.parentElement.querySelector('.collapse-icon').classList.add('collapsed');
            });
        }
                
        // 二级功能：按严重程度筛选问题
        let currentSeverityFilter = 'all';
                
        function filterIssuesBySeverity(severity) {
            currentSeverityFilter = severity;
                    
            // 更新仪表盘样式
            document.querySelectorAll('.dashboard-item').forEach((item, index) => {
                const severities = ['all', 'critical', 'major', 'minor', 'suggestion'];
                const itemSeverity = severities[index];
                        
                if (severity === itemSeverity) {
                    item.style.boxShadow = '0 4px 12px rgba(3, 102, 214, 0.3)';
                    item.style.borderColor = 'rgba(3, 102, 214, 0.5)';
                } else {
                    item.style.boxShadow = 'none';
                    item.style.borderColor = 'transparent';
                }
            });
                    
            // 筛选问题
            const allIssueItems = document.querySelectorAll('.issue-item');
            allIssueItems.forEach(item => {
                if (severity === 'all') {
                    item.style.display = 'block';
                } else {
                    const badge = item.querySelector('.severity-badge');
                    if (badge) {
                        const classMatch = badge.className.match(/badge-(\w+)/);
                        if (classMatch && classMatch[1] === severity) {
                            item.style.display = 'block';
                        } else {
                            item.style.display = 'none';
                        }
                    }
                }
            });
                    
            // 也筛选严重问题卡片
            const criticalCards = document.querySelectorAll('.critical-issue-card');
            criticalCards.forEach(card => {
                if (severity === 'all' || severity === 'critical') {
                    card.style.display = 'block';
                } else {
                    card.style.display = 'none';
                }
            });
        }
        
        // 代码段落切换函数
        function toggleCodeSnippet(header) {
            const content = header.nextElementSibling;
            const toggle = header.querySelector('.code-snippet-toggle');
            
            if (content) {
                if (content.classList.contains('collapsed')) {
                    content.classList.remove('collapsed');
                    if (toggle) toggle.classList.remove('collapsed');
                } else {
                    content.classList.add('collapsed');
                    if (toggle) toggle.classList.add('collapsed');
                }
            }
        }
    </script>
</body>
</html>
"""
