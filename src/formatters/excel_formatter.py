"""Excel格式化器"""
from typing import Dict, Any
from .base_formatter import BaseFormatter
from ..utils.data_processor import DataProcessor

# 尝试导入openpyxl
try:
    from openpyxl import Workbook as OpenpyxlWorkbook
    from openpyxl.styles import PatternFill as OpenpyxlPatternFill
    from openpyxl.styles import Font as OpenpyxlFont
    from openpyxl.styles import Alignment as OpenpyxlAlignment
    from openpyxl.styles import Border as OpenpyxlBorder
    from openpyxl.styles import Side as OpenpyxlSide
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 严重程度标签
SEVERITY_LABELS = {
    'critical': '严重',
    'major': '主要',
    'minor': '次要',
    'suggestion': '建议'
}


class ExcelFormatter(BaseFormatter):
    """Excel报告格式化器"""
    
    def __init__(self, output_dir: str = "./reports"):
        super().__init__(output_dir)
        # 不在初始化时检查，而是在format时检查
        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl 库未安装，Excel格式化器将不可用")
    
    def format(self, review_data: Dict[str, Any], **kwargs) -> str:
        """格式化为Excel报告
        
        注意：Excel格式化器直接保存文件，返回文件路径
        
        Args:
            review_data: 评审数据
            **kwargs: 可选参数
                - filepath: 保存路径（必需）
            
        Returns:
            文件保存路径
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 库未安装，请运行: pip install openpyxl")
        
        filepath = kwargs.get('filepath')
        if not filepath:
            raise ValueError("Excel formatter requires 'filepath' parameter")
        
        # 验证数据
        if not self.validate_data(review_data):
            raise ValueError("Invalid review data")
        
        # 预处理数据
        review_data = self.pre_process(review_data)
        
        # 创建工作簿
        wb = OpenpyxlWorkbook()  # type: ignore
        
        # 删除默认空白Sheet
        if wb.sheetnames and wb.sheetnames[0] == 'Sheet':
            wb.remove(wb[wb.sheetnames[0]])  # type: ignore
        
        # 定义样式
        header_fill = OpenpyxlPatternFill(start_color="0366D6", end_color="0366D6", fill_type="solid")  # type: ignore
        header_font = OpenpyxlFont(bold=True, color="FFFFFF", size=11)  # type: ignore
        critical_fill = OpenpyxlPatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")  # type: ignore
        major_fill = OpenpyxlPatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")  # type: ignore
        minor_fill = OpenpyxlPatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # type: ignore
        center_align = OpenpyxlAlignment(horizontal="center", vertical="center", wrap_text=True)  # type: ignore
        left_align = OpenpyxlAlignment(horizontal="left", vertical="top", wrap_text=True)  # type: ignore
        border = OpenpyxlBorder(  # type: ignore
            left=OpenpyxlSide(style='thin'),  # type: ignore
            right=OpenpyxlSide(style='thin'),  # type: ignore
            top=OpenpyxlSide(style='thin'),  # type: ignore
            bottom=OpenpyxlSide(style='thin')  # type: ignore
        )
        
        # 创建概览页
        self._create_overview_sheet(wb, review_data)
        
        # 创建问题详情页
        self._create_issues_sheet(wb, review_data, header_fill, header_font,
                                  critical_fill, major_fill, minor_fill,
                                  center_align, left_align, border)
        
        # 保存文件
        wb.save(filepath)  # type: ignore
        
        return filepath
    
    def _create_overview_sheet(self, wb, review_data: Dict[str, Any]) -> None:
        """创建概览页"""
        ws = wb.create_sheet("概览")
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
        row = 1
        ws[f'A{row}'] = "代码评审报告"
        ws[f'A{row}'].font = OpenpyxlFont(size=14, bold=True)  # type: ignore
        ws.merge_cells(f'A{row}:B{row}')
        row += 2
        
        # 基本信息
        metadata = review_data['metadata']
        info_items = [
            ("评审分支", metadata['source_branch']),
            ("基准分支", metadata['target_branch']),
            ("评审时间", metadata['review_time']),
            ("评审耗时", f"{metadata['duration_seconds']:.2f} 秒"),
            ("提交数量", str(metadata['total_commits'])),
            ("文件变更", str(metadata['total_files_changed'])),
        ]
        
        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = OpenpyxlFont(bold=True)  # type: ignore
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        # 统计信息
        stats = review_data['statistics']
        ws[f'A{row}'] = "问题统计"
        ws[f'A{row}'].font = OpenpyxlFont(size=12, bold=True)  # type: ignore
        row += 1
        
        stat_items = [
            ("总问题数", str(stats['total_issues'])),
            ("严重问题", str(stats['by_severity']['critical'])),
            ("主要问题", str(stats['by_severity']['major'])),
            ("次要问题", str(stats['by_severity']['minor'])),
            ("建议", str(stats['by_severity']['suggestion'])),
            ("代码增加", f"+{stats['total_additions']}"),
            ("代码删除", f"-{stats['total_deletions']}"),
        ]
        
        for label, value in stat_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = OpenpyxlFont(bold=True)  # type: ignore
            ws[f'B{row}'] = value
            row += 1
    
    def _create_issues_sheet(self, wb, review_data: Dict[str, Any],
                            header_fill, header_font, critical_fill,
                            major_fill, minor_fill, center_align,
                            left_align, border) -> None:
        """创建问题详情页"""
        ws_issues = wb.create_sheet("问题详情")
        ws_issues.column_dimensions['A'].width = 15
        ws_issues.column_dimensions['B'].width = 15
        ws_issues.column_dimensions['C'].width = 30
        ws_issues.column_dimensions['D'].width = 15
        ws_issues.column_dimensions['E'].width = 15
        ws_issues.column_dimensions['F'].width = 40
        ws_issues.column_dimensions['G'].width = 60
        ws_issues.column_dimensions['H'].width = 50
        
        # 表头
        headers = ["严重程度", "提交人", "文件", "行号", "方法", "问题描述", "改进建议", "问题代码"]
        for col, header in enumerate(headers, 1):
            cell = ws_issues.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row = 2
        
        # 收集所有问题
        all_issues = []
        for file_review in review_data.get('file_reviews', []):
            for issue in file_review.get('issues', []):
                issue_copy = issue.copy()
                issue_copy['file_path'] = file_review['file_path']
                all_issues.append(issue_copy)
        
        # 按严重程度排序
        all_issues = DataProcessor.sort_issues_by_severity(all_issues)
        
        # 填充数据
        for issue in all_issues:
            severity = issue['severity']
            
            ws_issues.cell(row=row, column=1).value = SEVERITY_LABELS.get(severity, severity)
            ws_issues.cell(row=row, column=2).value = issue.get('author', 'Unknown')
            ws_issues.cell(row=row, column=3).value = issue.get('file_path', 'N/A')
            ws_issues.cell(row=row, column=4).value = issue.get('line', 'N/A')
            ws_issues.cell(row=row, column=5).value = issue.get('method', 'N/A')
            ws_issues.cell(row=row, column=6).value = issue.get('description', '')
            ws_issues.cell(row=row, column=7).value = issue.get('suggestion', '')
            
            # 提取代码段落
            code_snippet_text = ''
            if issue.get('code_snippet'):
                snippet = issue['code_snippet']
                lines = snippet.get('lines', [])
                if lines:
                    code_lines = []
                    for line_obj in lines:
                        line_num = line_obj.get('line_num', '')
                        content = line_obj.get('content', '')
                        prefix = '+' if line_obj.get('type') == 'added' else '-' if line_obj.get('type') == 'deleted' else ' '
                        code_lines.append(f"{prefix} {line_num}: {content}")
                    code_snippet_text = '\n'.join(code_lines)
            
            ws_issues.cell(row=row, column=8).value = code_snippet_text if code_snippet_text else 'N/A'
            
            # 应用样式和边框
            for col in range(1, 9):
                cell = ws_issues.cell(row=row, column=col)
                cell.border = border
                cell.alignment = left_align
                
                # 根据严重程度填充背景色
                if severity == 'critical':
                    cell.fill = critical_fill
                elif severity == 'major':
                    cell.fill = major_fill
                elif severity == 'minor':
                    cell.fill = minor_fill
            
            row += 1
    

    
    def get_file_extension(self) -> str:
        """获取文件扩展名"""
        return ".xlsx"
